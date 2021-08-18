# Author: Kevin Zong
# Last Modified: August 17th, 2021
# Class representing Growing CTE/KDE data

from abc import ABC, abstractclassmethod, abstractmethod
from typing import List, Type
import os, sys

dir_path = os.path.dirname(os.path.realpath(__file__))
parent_dir_path = os.path.abspath(os.path.join(dir_path, os.pardir))
sys.path.insert(0, parent_dir_path)

from JSONDeserialization.epcis_event import (
    QuantityElement,
    QuantityEvent,
    URI,
    CommonEvent,
    AggregationEvent,
    EPCISEvent,
    ObjectEvent,
    TransactionEvent,
    TransformationEvent,
)
from epcis_cte_transformation.cte import CTEBase
import json
import datetime
from tools.serializer import jsonid
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook, load_workbook
from tools.serializer import JSONValueProvider, jsonid, map_from_json, map_to_json

class GrowingCTE:
    """
    Provides a class for the Growing CTE and defines its KDEs

    KDEs:

            Traceability Lot Code : List
            Growing Coordinates : str

            ---Sprout Grower Specification---
            Location of Seed Grower : str
            Seed Lot Code of Grower : str
            Date of Seed Harvesting : datetime
            Location of Seed Conditioner : str
            Seed Lot Code of Conditioner : str
            Date of Conditioning or Processing : datetime
            Location of Seed Packinghouse : List[str]
            Seed Lot Codes of Packinghouses : List[str]
            Date of Packing and Repacking : List[datetime]
            Location of Seed Supplier : str
            Description of Seeds : str
            Seed Lot Code of Seed Supplier : str
            Seed Receipt Date : datetime
            Sprout Traceability Lot Code(s) : str
            Date(s) of Production for each Seed Lot Code : List[datetime]
    """

    def __init__(self):
        self._traceability_lot_code = []
        self._growing_location = ""
        self._seed_grower_location = ""
        self._seed_grower_lot_code = ""
        self._seed_harvest_date = datetime.datetime(1,1,1)
        self._seed_conditioner_location = ""
        self._seed_conditioner_lot_code = ""
        self._seed_conditioning_date = datetime.datetime(1,1,1)
        self._seed_packinghouse_location = []
        self._seed_packinghouse_lot_code = []
        self._seed_packinghouse_date = []
        self._seed_supplier_location = ""
        self._seed_description = ""
        self._seed_supplier_lot_code = ""
        self._seed_receipt_date = datetime.datetime(1,1,1)
        self._sprout_traceability_lot_code = ""
        self._seed_lot_code_production_dates = []

    @classmethod 
    def new_from_data(cls, data: dict):
        pass

    @classmethod 
    def new_from_epcis(cls, event: EPCISEvent):
        output = cls()

        if(issubclass(type(event), CommonEvent)):
            try:
                output.growing_location = event.business_location.value
            except ValueError:
                output.growing_location = ""
        if isinstance(event, ObjectEvent):
            for epc in event.epc_list:
                output.traceability_lot_code.append(epc.value)
        
        return output

    @classmethod 
    def new_from_json(cls, json_data: str):
        data = json.loads(json_data)
        return cls.new_from_data(data)

    @classmethod 
    def new_from_excel(cls, excel_data: str):
        pass

    @classmethod 
    def save_to_excel(self):
        pass

    @classmethod 
    def new_from_csv(cls, csv_lines: "list[str]"):
        pass

    @property
    @jsonid("traceabilityLotCode")
    def traceability_lot_code(self) -> List:
        return self._traceability_lot_code

    @traceability_lot_code.setter
    def traceability_lot_code(self, value: List):
        self._traceability_lot_code = value

    @property
    @jsonid("growingLocation")
    def growing_location(self) -> str:
        return self._growing_location

    @growing_location.setter
    def growing_location(self, value: str):
        self._growing_location = value  

    @property
    @jsonid("seedGrowerLocation")
    def seed_grower_location(self) -> str:
        return self._seed_grower_location

    @seed_grower_location.setter
    def seed_grower_location(self, value: str):
        self._seed_grower_location = value

    @property
    @jsonid("seedGrowerLotCode")
    def seed_grower_lot_code(self) -> str:
        return self._seed_grower_lot_code

    @seed_grower_lot_code.setter
    def seed_grower_lot_code(self, value: str):
        self._seed_grower_lot_code = value 

    @property
    @jsonid("seedHarvestDate")
    def seed_harvest_date(self) -> datetime.datetime:
        return self._seed_harvest_date

    @seed_harvest_date.setter
    def seed_harvest_date(self, value: datetime.datetime):
        self._seed_harvest_date = value 

    @property
    @jsonid("seedConditionerLocation")
    def seed_conditioner_location(self) -> str:
        return self._seed_conditioner_location

    @seed_conditioner_location.setter
    def seed_conditioner_location(self, value: str):
        self._seed_conditioner_location = value 

    @property
    @jsonid("seedConditionerLotCode")
    def seed_conditioner_lot_code(self) -> str:
        return self._seed_conditioner_lot_code

    @seed_conditioner_lot_code.setter
    def seed_conditioner_lot_code(self, value: str):
        self._seed_conditioner_lot_code = value 

    @property
    @jsonid("seedConditioningDate")
    def seed_conditioning_date(self) -> datetime.datetime:
        return self._seed_conditioning_date

    @seed_conditioning_date.setter
    def seed_conditioning_date(self, value: datetime.datetime):
        self._seed_conditioning_date = value 

    @property
    @jsonid("seedPackinghouseLocation")
    def seed_packinghouse_location(self) -> List:
        return self._seed_packinghouse_location

    @seed_packinghouse_location.setter
    def seed_packinghouse_location(self, value: List):
        self._seed_packinghouse_location = value 

    @property
    @jsonid("seedPackinghouseLotCode")
    def seed_packinghouse_lot_code(self) -> List:
        return self._seed_packinghouse_lot_code

    @seed_packinghouse_lot_code.setter
    def seed_packinghouse_lot_code(self, value: List):
        self._seed_packinghouse_lot_code = value 

    @property
    @jsonid("seedPackinghouseDate")
    def seed_packinghouse_date(self) -> List[datetime.datetime]:
        return self._seed_packinghouse_date

    @seed_packinghouse_date.setter
    def seed_packinghouse_date(self, value: List[datetime.datetime]):
        self._seed_packinghouse_date = value 

    @property
    @jsonid("seedSupplierLocation")
    def seed_supplier_location(self) -> str:
        return self._seed_supplier_location

    @seed_supplier_location.setter
    def seed_supplier_location(self, value: str):
        self._seed_supplier_location = value 

    @property
    @jsonid("seedDescription")
    def seed_description(self) -> str:
        return self._seed_description

    @seed_description.setter
    def seed_description(self, value: str):
        self._seed_description = value 

    @property
    @jsonid("seedSupplierLotCode")
    def seed_supplier_lot_code(self) -> str:
        return self._seed_supplier_lot_code

    @seed_supplier_lot_code.setter
    def seed_supplier_lot_code(self, value: str):
        self._seed_supplier_lot_code = value 

    @property
    @jsonid("seedReceiptData")
    def seed_receipt_date(self) -> datetime.datetime:
        return self._seed_receipt_date

    @seed_receipt_date.setter
    def seed_receipt_date(self, value: datetime.datetime):
        self._seed_receipt_date = value 

    @property
    @jsonid("sproutTraceabilityLotCode")
    def sprout_traceability_lot_code(self) -> str:
        return self._sprout_traceability_lot_code

    @sprout_traceability_lot_code.setter
    def sprout_traceability_lot_code(self, value: str):
        self._sprout_traceability_lot_code = value 

    @property
    @jsonid("seedLotCodeProductionDates")
    def seed_lot_code_production_dates(self) -> List:
        return self._seed_lot_code_production_dates

    @seed_lot_code_production_dates.setter
    def seed_lot_code_production_dates(self, value: List):
        self._seed_lot_code_production_dates = value          

    @classmethod
    def output_json(self) -> str:
        data = map_to_json(self)
        return json.dumps(data)

    def output_xlsx(self, sheet: Worksheet, row) -> str:
        """
        Create an excel spreadsheet and output the contents to an XML string
        """

        kde_ids = [
            "Traceability Lot Code",
            "Growing Area"
        ]

        kde_values = [
            self.traceability_lot_code,
            self.growing_location
        ]
        if row == 1:
            for i in range(1, 3):
                cell = sheet.cell(row=row, column=i)
                cell.value = kde_ids[i - 1]

        for i in range(1, 3):
            cell = sheet.cell(row=row + 1, column=i)
            tmpval = kde_values[i - 1]
            if isinstance(tmpval, list):
                tmpval = ", ".join(tmpval)
            cell.value = tmpval


        sheet.row_dimensions[1].height = 30
        sheet.row_dimensions[2].height = 30
        sheet.column_dimensions["A"].width = 40
        sheet.column_dimensions["B"].width = 40

    @classmethod 
    def save_as_xlsx(self, filename: str):
        pass
        # code here