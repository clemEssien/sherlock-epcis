from abc import ABC, abstractclassmethod, abstractmethod
from typing import List, Type
import os, sys
import datetime
import json

dir_path = os.path.dirname(os.path.realpath(__file__))
parent_dir_path = os.path.abspath(os.path.join(dir_path, os.pardir))
sys.path.insert(0, parent_dir_path)

from JSONDeserialization.epcis_event import (
    QuantityElement,
    QuantityEvent,
    URI,
    CommonEvent,
    AggregationEvent,
    EPCISEvent,
    ObjectEvent,
    TransactionEvent,
    TransformationEvent,
)

from epcis_cte_transformation.cte import CTEBase

from openpyxl import Workbook, load_workbook
from tools.serializer import JSONValueProvider, jsonid, map_from_json, map_to_json


class TransformationCTE(CTEBase):
    """
    Provides a class for the Transformation CTE and defines its KDEs

    KDEs:
            Traceability Product : List
            Quantity of Input : List
            Quantity of Output : List
            Location of Transformation : str
            New Traceability Product : List
            Unit of Measure : List
    """

    def __init__(self):
        """Creates new Transformation CTE"""
        self._traceability_product = []
        self._quantity_of_input = []
        self._quantity_of_output = []
        self._location_of_transformation = ""
        self._new_traceability_product = []
        self._unit_of_measure = []

    @classmethod
    def new_from_data(cls, data: dict):
        pass

    @classmethod
    def new_from_epcis(cls, event: EPCISEvent):
        output = cls()

        # check if TransformationCTE must come from TransformationEvent
        # transEvent: TransformationEvent = event
        if isinstance(event, TransformationEvent):
            try:
                output.location_of_transformation = event.read_point.value
            except ValueError:
                output.location_of_transformation = ""
            for qe in event.input_quantity_list:
                output.quantity_of_input.append(qe.quantity)
                output.unit_of_measure.append(qe.uom)
            for qe in event.output_quantity_list:
                output.quantity_of_output.append(qe.quantity)
            for epc in event.input_epc_list:
                output.traceability_product.append(epc.value)
            for epc in event.output_epc_list:
                output.new_traceability_product.append(epc.value)
        elif isinstance(event, ObjectEvent):
            try:
                output.location_of_transformation = event.read_point.value
            except ValueError:
                output.location_of_transformation = ""
            for qe in event.quantity_list:
                output.quantity_of_input.append(qe.quantity)
                output.unit_of_measure.append(qe.uom)
            for epc in event.epc_list:
                output.traceability_product.append(epc.value)
        elif isinstance(event, AggregationEvent):
            try:
                output.location_of_transformation = event.read_point.value
            except ValueError:
                output.location_of_transformation = ""
            for qe in event.child_quantity_list:
                output.quantity_of_input.append(qe.quantity)
                output.unit_of_measure.append(qe.uom)
            for epc in event.child_epc_list:  # check if we should use this or parentID
                output.traceability_product.append(epc.value)
        elif isinstance(event, TransactionEvent):
            try:
                output.location_of_transformation = event.read_point.value
            except ValueError:
                output.location_of_transformation = ""
            for qe in event.quantity_list:
                output.quantity_of_input.append(qe.quantity)
                output.unit_of_measure.append(qe.uom)
            for epc in event.epc_list:  # check if we should use this or parentID
                output.traceability_product.append(epc.value)
        elif isinstance(event, CommonEvent):
            try:
                output.location_of_transformation = event.read_point.value
            except ValueError:
                output.location_of_transformation = ""
        return output

    @classmethod
    def new_from_json(cls, json_data: str):
        """
        Create a new CTE from JSON data
        """
        # data = json.loads(json_data)
        # return cls.new_from_data(data)
        output = cls()
        types = {
            output.traceability_product: list,
            output.quantity_of_input: list,
            output.quantity_of_output: list,
            output.location_of_transformation: str,
            output.new_traceability_product: list,
            output.unit_of_measure: list,
        }

        map_from_json(json_data, output, types)
        return output

    @classmethod
    def new_from_excel(cls, excel_data: str):
        pass

    def save_to_excel(self):
        pass

    @classmethod
    def new_from_csv(cls, csv_lines: "list[str]"):
        pass


    @property
    @jsonid("traceabilityProduct")
    def traceability_product(self) -> str:
        return self._traceability_product

    @traceability_product.setter
    def traceability_product(self, value: str) -> None:
        self._traceability_product = value



    @property
    @jsonid("inputQuantity")
    def quantity_of_input(self) -> str:
        return self._quantity_of_input

    @quantity_of_input.setter
    def quantity_of_input(self, value: str) -> None:
        self._quantity_of_input = value



    @property
    @jsonid("outputQuantity")
    def quantity_of_output(self) -> str:
        return self._quantity_of_output

    @quantity_of_output.setter
    def quantity_of_output(self, value: str) -> None:
        self._quantity_of_output = value



    @property
    @jsonid("transformationLocation")
    def location_of_transformation(self) -> str:
        return self._location_of_transformation

    @location_of_transformation.setter
    def location_of_transformation(self, value: str) -> None:
        self._location_of_transformation = value



    @property
    @jsonid("newTraceabilityProduct")
    def new_traceability_product(self) -> str:
        return self._new_traceability_product

    @new_traceability_product.setter
    def new_traceability_product(self, value: str) -> None:
        self._new_traceability_product = value



    @property
    @jsonid("unit")
    def unit_of_measure(self) -> str:
        return self._unit_of_measure

    @unit_of_measure.setter
    def unit_of_measure(self, value: str) -> None:
        self._unit_of_measure = value


    def output_json(self) -> str:
        """
        Returnns serialized JSON data
        """
        data = map_to_json(self)
        return json.dumps(data)

    def output_xlsx(self) -> str:
        """
        Create an excel spreadsheet
        """

        workbook = Workbook()
        sheet = workbook.active
        filename = "transformation_cte.xlsx"

        kde_ids = [
            "Traceability Product",
            "Quantity of Input",
            "Quantity of Output",
            "Location of Transformation",
            "New Traceability Product",
            "Unit of Measure",
        ]
        traceability_product_str = ", ".join(self.traceability_product)
        quantity_of_input_str = ", ".join(self.quantity_of_input)
        quantity_of_output_str = ", ".join(self.quantity_of_output)
        new_traceability_product_str = ", ".join(self.new_traceability_product)
        uom_str = ", ".join(self.unit_of_measure)
        kde_values = [
            traceability_product_str,
            quantity_of_input_str,
            quantity_of_output_str,
            self.location_of_transformation,
            new_traceability_product_str,
            uom_str,
        ]

        for i in range(1, 7):
            cell = sheet.cell(row=1, column=i)
            cell.value = kde_ids[i - 1]

        for i in range(1, 7):
            cell = sheet.cell(row=2, column=i)
            cell.value = kde_values[i - 1]

        sheet.row_dimensions[1].height = 30
        sheet.row_dimensions[2].height = 30
        sheet.column_dimensions["A"].width = 40
        sheet.column_dimensions["B"].width = 40
        sheet.column_dimensions["C"].width = 40
        sheet.column_dimensions["D"].width = 40
        sheet.column_dimensions["E"].width = 40
        sheet.column_dimensions["F"].width = 40
        # /var/src/documents/<companyid>/<userid>/<cte types>/<name/id>_<timestamp>.xlsx
        # Unknown: companyID, userID, CTETypes, name/id
        # workbook.save('/var/src/documents/' + filename + " " + datetime.datetime.now + '.xlsx')
        workbook.save(filename)
        return filename

    def save_as_xlsx(self, filename: str):
        # will filename include .xlsx extentsion?
        workbook = load_workbook(filename)
        workbook.save(filename)
