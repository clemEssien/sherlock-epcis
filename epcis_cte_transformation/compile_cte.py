
from tools.serializer import jsonid, map_from_json, map_to_json
import json

from epcis_cte_transformation.ftl_food import FTLFood
from epcis_cte_transformation.location_master import LocationMaster

from epcis_cte_transformation.cte import CTEBase
from epcis_cte_transformation.creation_cte import CreationCTE
from epcis_cte_transformation.growing_cte import GrowingCTE
from epcis_cte_transformation.receiving_cte import ReceivingCTE
from epcis_cte_transformation.shipping_cte import ShippingCTE
from epcis_cte_transformation.transformation_cte import TransformationCTE

from abc import ABC, abstractclassmethod, abstractmethod
from typing import List
import os, sys
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

def create_filename():
    fn = "/var/src/uploads/" + str(datetime.now().isoformat("T").replace("+00:00", "") + "Z") + ".xlsx"
    
    
    pass

def compile_ctes(ctelist: "list[CTEBase]"):
    names = {
            "shipping": "Shipping KDEs",
            "growing": "Growing KDEs",
            "growing_sprouts": "Growing KDEs (If sprouts)",
            "receiving": "Receiving KDEs",
            "receiving_first": "Receiving KDEs (If first receiver except for seafood)",
            "receiving_first_seafood": "Receiving KDEs (If first receiver of seafood)",
            "transformation": "Transformation KDEs",
            "creation": "Creation KDEs",
            "ftlfoods": "FTL Foods",
            "location": "Location Master"
        }

    filename = create_filename()    
    sheets: "dict[str, Worksheet]" = {}
    
    workbook = Workbook()
    sheet = workbook.active    
    sheets["ftlfoods"] = sheet
    sheet.title = "FTL Foods"
    
    sheet = workbook.create_sheet("Location Master")
    sheets["location"] = sheet
    
    counts: "dict[str, int]" = {}
    
    for item in ctelist:
                
        cls: str = None

        if isinstance(item, CreationCTE):
            cls = "creation"
        elif isinstance(item, GrowingCTE):
            cls = "growing"
        elif isinstance(item, ReceivingCTE):
            cls = "receiving"
        elif isinstance(item, ShippingCTE):
            cls = "shipping"
        elif isinstance(item, TransformationCTE):
            cls = "transformation"
        elif isinstance(item, LocationMaster):
            cls = "location"
        elif isinstance(item, FTLFood):
            cls = "ftlfoods"

        if not cls in counts:
            counts[cls] = 1
        
        row = counts[cls] 
        counts[cls] = row + 1
        
        if not cls in sheets:
            sheets[cls] = workbook.create_sheet(names[cls])
        
        sheet = sheets[cls]
        item.output_xlsx(sheet, row)

    workbook.save(filename)
    return filename



    