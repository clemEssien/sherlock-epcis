from abc import ABC, abstractclassmethod, abstractmethod
from typing import List
from epcis_cte_transformation.cte import CTEBase
import os, sys
import datetime
import json

dir_path = os.path.dirname(os.path.realpath(__file__))
parent_dir_path = os.path.abspath(os.path.join(dir_path, os.pardir))
sys.path.insert(0, parent_dir_path)
from JSONDeserialization.epcis_event import (
    QuantityElement,
    QuantityEvent,
    URI,
    CommonEvent,
    AggregationEvent,
    EPCISEvent,
    ObjectEvent,
    TransactionEvent,
    TransformationEvent,
)

from openpyxl import Workbook, load_workbook
from tools.serializer import JSONValueProvider, jsonid, map_from_json, map_to_json


class CreationCTE(CTEBase):

    """
    Provides a class for the Creation CTE and defines its KDEs

    KDEs:
            Traceability Product : List
            Creation Completion Date : datetime
            Location Where Food Was Created : str
            Quantity : List
            Unit of Measure : List
    """

    def __init__(self) -> None:
        """Creates new Shipping CTE"""
        self._traceability_product = []
        self._creation_completion_date: datetime.datetime(1, 1, 1)
        self._loaction_where_food_was_created = ""
        self._quantity = []
        self._unit_of_measure = []

    @classmethod
    def new_from_data(cls, data: dict):
        pass

    @classmethod
    def new_from_epcis(cls, event: EPCISEvent):
        """
        Create a new CTE from an EPCIS event
        """

        output = cls()
        if isinstance(event, ObjectEvent):
            try:
                output.creation_completion_date = event.event_time_local
            except ValueError:
                output.creation_completion_date = ""
            try:
                output.location_where_food_was_created = event.read_point.value
            except ValueError:
                output.location_where_food_was_created = ""
            for epc in event.epc_list:
                output.traceability_product.append(epc.value)
            for qe in event.quantity_list:
                output.quantity.append(str(qe.quantity))
                output.unit_of_measure.append(qe.uom)
        elif isinstance(event, AggregationEvent):
            try:
                output.creation_completion_date = event.event_time_local
            except ValueError:
                output.creation_completion_date = ""
            try:
                output.location_where_food_was_created = event.read_point.value
            except ValueError:
                output.location_where_food_was_created = ""
            for epc in event.child_epc_list:
                output.traceability_product.append(epc.value)
            for qe in event.child_quantity_list:
                output.quantity.append(qe.quantity)
                output.unit_of_measure.append(qe.uom)
        elif isinstance(event, TransactionEvent):
            try:
                output.creation_completion_date = event.event_time_local
            except ValueError:
                output.creation_completion_date = ""
            try:
                output.location_where_food_was_created = event.read_point.value
            except ValueError:
                output.location_where_food_was_created = ""
            for epc in event.epc_list:
                output.traceability_product.append(epc.value)
            for qe in event.quantity_list:
                output.quantity.append(qe.quantity)
                output.unit_of_measure.append(qe.uom)
        elif isinstance(event, TransformationEvent):
            try:
                output.creation_completion_date = event.event_time_local
            except ValueError:
                output.creation_completion_date = ""
            try:
                output.location_where_food_was_created = event.read_point.value
            except ValueError:
                output.location_where_food_was_created = ""
            for epc in event.input_epc_list:
                output.traceability_product.append(epc.value)
            for qe in event.input_quantity_list:
                output.quantity.append(qe.quantity)
                output.unit_of_measure.append(qe.uom)
        elif isinstance(event, CommonEvent):
            try:
                output.creation_completion_date = event.event_time_local
            except ValueError:
                output.creation_completion_date = ""
            try:
                output.location_where_food_was_created = event.read_point.value
            except ValueError:
                output.location_where_food_was_created = ""
        return output

    @classmethod
    def new_from_json(cls, json_data: str):
        """
        Create a new CTE from JSON data
        """
        output = cls()
        types = {
            output.traceability_product: list,
            output.creation_completion_date: datetime.datetime,
            output.location_where_food_was_created: str,
            output.quantity: list,
            output.unit_of_measure: list,
        }

        map_from_json(json_data, output, types)
        return output

    def new_from_csv(cls, csv_lines: "list[str]"):
        pass

    def new_from_excel(cls, excel_data: str):
        pass

    def save_to_excel(self):
        pass

    @property
    @jsonid("traceabilityProduct")
    def traceability_product(self) -> List:
        return self._traceability_product

    @traceability_product.setter
    def traceability_product(self, value: List):
        self._traceability_product = value

    @property
    @jsonid("creationCompletionDate")
    def creation_completion_date(self) -> datetime.datetime:
        return self._creation_completion_date

    @creation_completion_date.setter
    def creation_completion_date(self, value: datetime.datetime):
        self._creation_completion_date = value

    @property
    @jsonid("locationWhereFoodWasCreated")
    def location_where_food_was_created(self) -> str:
        return self._location_where_food_was_created

    @location_where_food_was_created.setter
    def location_where_food_was_created(self, value: str):
        self._location_where_food_was_created = value

    @property
    @jsonid("quantity")
    def quantity(self) -> List:
        return self._quantity

    @quantity.setter
    def quantity(self, value: List):
        self._quantity = value

    @property
    @jsonid("unit")
    def unit_of_measure(self) -> List:
        return self._unit_of_measure

    @unit_of_measure.setter
    def unit_of_measure(self, value: List):
        self._unit_of_measure = value

    def output_json(self):
        """
        Returnns serialized JSON data
        """
        data = map_to_json(self)
        return json.dumps(data)

    def output_xlsx(self) -> str:
        """
        Create an excel spreadsheet
        """

        workbook = Workbook()
        sheet = workbook.active
        filename = "creation_cte.xlsx"
        kde_ids = [
            "Traceability Product",
            "Creation Completion Date",
            "Location Where Food Was Created",
            "Quantity",
            "Unit of Measure",
        ]
        traceability_product_str = ", ".join(self.traceability_product)
        quantity_str = ", ".join(self.quantity)
        uom_str = ", ".join(self.unit_of_measure)
        kde_values = [
            traceability_product_str,
            str(self.creation_completion_date),
            self.location_where_food_was_created,
            quantity_str,
            uom_str,
        ]
        for i in range(1, 6):
            cell = sheet.cell(row=1, column=i)
            cell.value = kde_ids[i - 1]

        for i in range(1, 6):
            cell = sheet.cell(row=2, column=i)
            cell.value = kde_values[i - 1]

        sheet.row_dimensions[1].height = 30
        sheet.row_dimensions[2].height = 30
        sheet.column_dimensions["A"].width = 40
        sheet.column_dimensions["B"].width = 40
        sheet.column_dimensions["C"].width = 40
        sheet.column_dimensions["D"].width = 40
        sheet.column_dimensions["E"].width = 40
        # /var/src/documents/<companyid>/<userid>/<cte types>/<name/id>_<timestamp>.xlsx
        # Unknown: companyID, userID, CTETypes, name/id
        # workbook.save('/var/src/documents/' + filename + " " + datetime.datetime.now + '.xlsx')
        workbook.save(filename)
        return filename

    def save_as_xlsx(self, filename: str):
        # will filename include .xlsx extension?
        workbook = load_workbook(filename)
        workbook.save(filename)
