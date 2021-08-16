from abc import ABC, abstractclassmethod, abstractmethod
from typing import List

from openpyxl.worksheet.worksheet import Worksheet
from epcis_cte_transformation.cte import CTEBase
import os, sys
import datetime
import json

from tools.serializer import jsonid

dir_path = os.path.dirname(os.path.realpath(__file__))
parent_dir_path = os.path.abspath(os.path.join(dir_path, os.pardir))
sys.path.insert(0, parent_dir_path)
from JSONDeserialization.epcis_event import (
    QuantityElement,
    QuantityEvent,
    URI,
    CommonEvent,
    AggregationEvent,
    EPCISEvent,
    ObjectEvent,
    TransactionEvent,
    TransformationEvent,
)

from openpyxl import Workbook, load_workbook
from tools.serializer import JSONValueProvider, jsonid, map_from_json, map_to_json


class ShippingCTE(CTEBase):

    """
    Provides a class for the Shipping CTE and defines its KDEs

    KDEs:
            Traceability Lot Code : List
            Entry Number : str
            Quantity : List
            Unit of Measure : List
            Traceability Product : List
            Location of Traceability Lot Code Generator : str
            Location of Recipient : str
            Location of Source of Shipment : str
            Transporter Name: str
            Lot Code PoC Name: str
            Lot Code PoC Phone: str
            Lot Code PoC Email: str
            Shipment Datetime: datetime

    """

    def __init__(self) -> None:
        """Creates new Shipping CTE"""
        self._traceability_lot_code = []
        self._entry_number = ""
        self._quantity = []
        self._unit_of_measure = []
        self._traceability_product = []
        self._location_of_traceability_lot_code_generator = ""
        self._location_of_recipient = ""
        self._location_of_source_of_shipment = ""
        self._transporter_name = ""
        self._lot_code_poc_name = ""
        self._lot_code_poc_phone = ""
        self._lot_code_poc_email = ""
        self._shipment_datetime: datetime.datetime(1, 1, 1)

    @classmethod
    def new_from_data(cls, data: dict):
        pass
    @classmethod
    def new_from_epcis(cls, event: EPCISEvent):
        output = cls()
        if isinstance(event, ObjectEvent):
            try:
                output.location_of_recipient = event.business_location.value
            except ValueError:
                output.location_of_recipient = ""
            try:
                output.location_of_source_of_shipment = event.read_point.value
            except ValueError:
                output.location_of_source_of_shipment = ""
            try:
                output.shipment_datetime = event.event_time_local
            except ValueError:
                output.shipment_datetime = ""
            for qe in event.quantity_list:
                output.quantity.append(qe.quantity)
                output.unit_of_measure.append(qe.uom)
                output.traceability_lot_code.append(qe.epc_class.value)
            for epc in event.epc_list:  # check what should be traceability lot code
                output.traceability_product.append(epc.value)
        elif isinstance(event, AggregationEvent):
            try:
                output.location_of_recipient = event.business_location.value
            except ValueError:
                output.location_of_recipient = ""
            try:
                output.location_of_source_of_shipment = event.read_point.value
            except ValueError:
                output.location_of_source_of_shipment = ""
            try:
                output.shipment_datetime = event.event_time_local
            except ValueError:
                output.shipment_datetime = ""
            for qe in event.child_quantity_list:
                output.quantity.append(qe.quantity)
                output.unit_of_measure.append(qe.uom)
                output.traceability_lot_code.append(qe.epc_class.value)
            for epc in event.child_epc_list:
                output.traceability_product.append(epc.value)
        elif isinstance(event, TransactionEvent):
            try:
                output.location_of_recipient = event.business_location.value
            except ValueError:
                output.location_of_recipient = ""
            try:
                output.location_of_source_of_shipment = event.read_point.value
            except ValueError:
                output.location_of_source_of_shipment = ""
            try:
                output.shipment_datetime = event.event_time_local
            except ValueError:
                output.shipment_datetime = ""
            for qe in event.quantity_list:
                output.quantity.append(qe.quantity)
                output.unit_of_measure.append(qe.uom)
                output.traceability_lot_code.append(qe.epc_class.value)
            for epc in event.epc_list:
                output.traceability_product.append(epc.value)
        elif isinstance(event, TransformationEvent):
            try:
                output.location_of_recipient = event.business_location.value
            except ValueError:
                output.location_of_recipient = ""
            try:
                output.location_of_source_of_shipment = event.read_point.value
            except ValueError:
                output.location_of_source_of_shipment = ""
            try:
                output.shipment_datetime = event.event_time_local
            except ValueError:
                output.shipment_datetime = ""
            for qe in event.input_quantity_list:
                output.quantity.append(qe.quantity)
                output.unit_of_measure.append(qe.uom)
                output.traceability_lot_code.append(qe.epc_class.value)
            for epc in event.input_epc_list:
                output.traceability_product.append(epc.value)
        elif isinstance(event, CommonEvent):
            try:
                output.location_of_recipient = event.business_location.value
            except ValueError:
                output.location_of_recipient = ""
            try:
                output.location_of_source_of_shipment = event.read_point.value
            except ValueError:
                output.location_of_source_of_shipment = ""
            try:
                output.shipment_datetime = event.event_time_local
            except ValueError:
                output.shipment_datetime = ""

        return output

    @classmethod
    def new_from_json(cls, json_data: str):
        """
        Create a new CTE from JSON data
        """
        output = cls()
        types = {
            output.traceability_lot_code: list,
            output.entry_number: str,
            output.quantity: list,
            output.unit_of_measure: list,
            output.traceability_product: list,
            output.location_of_traceability_lot_code_generator: str,
            output.location_of_recipient: str,
            output.location_of_source_of_shipment: str,
            output.transporter_name: str,
            output.lot_code_poc_name: str,
            output.lot_code_poc_phone: str,
            output.lot_code_poc_email: str,
            output.shipment_datetime: datetime.datetime,
        }

        map_from_json(json_data, output, types)
        return output

    @classmethod
    def new_from_csv(cls, csv_lines: "list[str]"):
        pass

    @classmethod
    def new_from_excel(cls, excel_data: str):
        pass

    def save_to_excel(self):
        pass

    #     transporterName: "",
    #     importEntryNumber: "",
    #     lotCode: "OL001",
    #     quantity: "100",
    #     traceabilityPid: "9991002100014",
    #     locationId: "",
    #     lotCodePocName: "",
    #     lotCodePocPhone: "",
    #     lotCodePocEmail: "",
    #     immediateRecipientLocationId: "9991000100023",
    #     shippedLocationId: "9991002100014",
    #     shipmentDateTime: "2021-07-15T00:00:00+02:00",

    @property
    @jsonid("lotCode")
    def traceability_lot_code(self) -> str:
        return self._traceability_lot_code

    @traceability_lot_code.setter
    def traceability_lot_code(self, value: str):
        self._traceability_lot_code = value

    @property
    @jsonid("importEntryNumber")
    def entry_number(self) -> str:
        return self._entry_number

    @entry_number.setter
    def entry_number(self, value: str):
        self._entry_number = value

    @property
    @jsonid("quantity")
    def quantity(self) -> List:
        return self._quantity

    @quantity.setter
    def quantity(self, value: List):
        self._quantity = value

    @property
    @jsonid("unit")
    def unit_of_measure(self) -> List:
        return self._unit_of_measure

    @unit_of_measure.setter
    def unit_of_measure(self, value: List):
        self._unit_of_measure = value

    @property
    @jsonid("traceabilityPid")
    def traceability_product(self) -> List:
        return self._traceability_product

    @traceability_product.setter
    def traceability_product(self, value: List):
        self._traceability_product = value

    @property
    @jsonid("locationId")
    def location_of_traceability_lot_code_generator(self) -> str:
        return self._location_of_traceability_lot_code_generator

    @location_of_traceability_lot_code_generator.setter
    def location_of_traceability_lot_code_generator(self, value: str):
        self._location_of_traceability_lot_code_generator = value

    @property
    @jsonid("immediateRecipientLocationId")
    def location_of_recipient(self) -> str:
        return self._location_of_recipient

    @location_of_recipient.setter
    def location_of_recipient(self, value: str):
        self._location_of_recipient = value

    @property
    @jsonid("shippedLocationId")
    def location_of_source_of_shipment(self) -> str:
        return self._location_of_source_of_shipment

    @location_of_source_of_shipment.setter
    def location_of_source_of_shipment(self, value: str):
        self._location_of_source_of_shipment = value

    @property
    @jsonid("transporterName")
    def transporter_name(self) -> str:
        return self._transporter_name

    @transporter_name.setter
    def transporter_name(self, value: str):
        self._transporter_name = value

    @property
    @jsonid("lotCodePocName")
    def lot_code_poc_name(self) -> str:
        return self._lot_code_poc_name

    @lot_code_poc_name.setter
    def lot_code_poc_name(self, value: str):
        self._lot_code_poc_name = value

    @property
    @jsonid("lotCodePocPhone")
    def lot_code_poc_phone(self) -> str:
        return self._lot_code_poc_phone

    @lot_code_poc_phone.setter
    def lot_code_poc_phone(self, value: str):
        self._lot_code_poc_phone = value

    @property
    @jsonid("lotCodePocEmail")
    def lot_code_poc_email(self) -> str:
        return self._lot_code_poc_email

    @lot_code_poc_email.setter
    def lot_code_poc_email(self, value: str) -> str:
        self._lot_code_poc_email = value

    @property
    @jsonid("shipmentDateTime")
    def shipment_datetime(self) -> datetime.datetime:
        return self._shipment_datetime

    @shipment_datetime.setter
    def shipment_datetime(self, value: str):
        self._shipment_datetime = value

    def output_json(self):
        """
        Returnns serialized JSON data
        """
        data = map_to_json(self)
        return json.dumps(data)

    def output_xlsx(self, sheet: Worksheet, row) -> str:
        """
        Create an excel spreadsheet and output the contents to an XML string
        """

        kde_ids = [
            "Traceability Lot Code",
            "Entry Number",
            "Quantity",
            "Unit of Measure",
            "Traceability Product",
            "Location of Traceability Lot Code Generator",
            "Location of Recipient",
            "Location of Source of Shipment",
            "Transporter Name",
            "Lot Code PoC Name",
            "Lot Code PoC Phone",
            "Lot Code PoC Email",
            "Shipment Datetime",
        ]

        traceability_lot_code_str = ", ".join(self.traceability_lot_code)
        quantity_str = ", ".join(self.quantity)
        uom_str = ", ".join(self.unit_of_measure)
        traceability_product_str = ", ".join(self.traceability_product)

        kde_values = [
            traceability_lot_code_str,
            self.entry_number,
            quantity_str,
            uom_str,
            traceability_product_str,
            self.location_of_traceability_lot_code_generator,
            self.location_of_recipient,
            self.location_of_source_of_shipment,
            self.transporter_name,
            self.lot_code_poc_name,
            self.lot_code_poc_phone,
            self.lot_code_poc_email,
            str(self.shipment_datetime),
        ]
        
        
        if row == 1:
            for i in range(1, 14):
                cell = sheet.cell(row=row, column=i)
                cell.value = kde_ids[i - 1]

        for i in range(1, 14):
            cell = sheet.cell(row=row+1, column=i)
            cell.value = kde_values[i - 1]

        sheet.row_dimensions[1].height = 30
        sheet.row_dimensions[2].height = 30
        sheet.column_dimensions["A"].width = 40
        sheet.column_dimensions["B"].width = 40
        sheet.column_dimensions["C"].width = 40
        sheet.column_dimensions["D"].width = 40
        sheet.column_dimensions["E"].width = 40
        sheet.column_dimensions["F"].width = 40
        sheet.column_dimensions["G"].width = 40
        sheet.column_dimensions["H"].width = 40
        sheet.column_dimensions["I"].width = 40
        sheet.column_dimensions["J"].width = 40
        sheet.column_dimensions["K"].width = 40
        sheet.column_dimensions["L"].width = 40
        sheet.column_dimensions["M"].width = 40
        # /var/src/documents/<companyid>/<userid>/<cte types>/<name/id>_<timestamp>.xlsx
        # Unknown: companyID, userID, CTETypes, name/id
        # workbook.save('/var/src/documents/' + filename + " " + datetime.datetime.now + '.xlsx')

    def save_as_xlsx(self, filename: str):
        # will filename include .xlsx extentsion?
        workbook = load_workbook(filename)
        workbook.save(filename)
